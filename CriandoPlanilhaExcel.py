#> > > > > Criando Planilha no Excel < < < < <

from random import randint
from openpyxl import Workbook

rand = (1, 100)
rows, cols = 10, 10

wb = Workbook()
ws = wb.active

for r in range(1, rows+1):
    for c in range(1, cols+1):
        ws.cell(row=r, column=c, value=randint(*rand))

ws.title = 'Aba_Teste'
wb.save('PlanilhaTeste.xlsx')

#> > > > > Criando Planilha no Excel < < < < <
'''
from colorama import Back, Fore, init
from openpyxl import load_workbook

workbook, worksheet = 'Teste.xlsx', 'Aba_1'
rows, cols = 10, 10

wb = load_workbook(workbook)
ws = wb[worksheet]

init()

for r in range(1, rows+1):
    print(
        Fore.CYAN if r % 2 else Fore.MAGENTA, ' | '.join([
            '{:5d}'.format(ws.cell(row=r, column=c).value)
                for c in range(1, cols+1)
        ]),Fore.RESET, sep=''
    )
'''